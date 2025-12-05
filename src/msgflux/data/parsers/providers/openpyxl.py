from typing import Dict, List, Optional, Tuple, Union

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from msgflux.data.parsers.base import BaseParser
from msgflux.data.parsers.registry import register_parser
from msgflux.data.parsers.response import ParserResponse
from msgflux.data.parsers.types import XlsxParser
from msgflux.dotdict import dotdict


@register_parser
class OpenPyxlXlsxParser(BaseParser, XlsxParser):
    """OpenPyxl-based XLSX Parser.

    Converts Excel spreadsheets to Markdown format and extracts embedded images.
    Uses openpyxl library for Excel processing.

    Features:
    - Multi-sheet support
    - Table extraction (converted to Markdown tables)
    - Image extraction from sheets
    - Merged cell handling
    - Sheet-level organization

    Example:
        >>> parser = Parser.xlsx("openpyxl")
        >>> response = parser("spreadsheet.xlsx")
        >>> print(response.data["text"])
        >>> print(response.data["images"])
    """

    provider = "openpyxl"

    def __init__(
        self,
        *,
        table_format: Optional[str] = "markdown",
        encode_images_base64: Optional[bool] = True,
    ):
        """Initialize OpenPyxl parser.

        Args:
            table_format:
                Format for table output. Options: "markdown" or "html".
            encode_images_base64:
                If True, encode images as base64 strings.
                If False, keep as raw bytes.
        """
        if load_workbook is None:
            raise ImportError(
                "`openpyxl` is not available. Install with `pip install openpyxl`"
            )

        self.table_format = table_format
        self.encode_images_base64 = encode_images_base64
        self._initialize()

    def _initialize(self):
        """Initialize parser state."""
        pass

    def __call__(self, data: Union[str, bytes], **kwargs) -> ParserResponse:
        """Parse an XLSX document.

        Args:
            data:
                XLSX file path, URL, or bytes.
            **kwargs:
                Additional parsing options (currently unused).

        Returns:
            ParserResponse containing:
            - text: Markdown-formatted content with tables
            - images: Dictionary of extracted images
            - metadata: Document metadata (num_sheets, sheet_names, etc.)

        Raises:
            FileNotFoundError:
                If file path doesn't exist.
            ValueError:
                If data type is not supported or file type is invalid.
        """
        # Validate file type if it's a path
        if isinstance(data, str) and not data.startswith(("http://", "https://")):
            self._validate_file_type(data, [".xlsx", ".xlsm"])

        # Parse the document
        result = self._parse(data)

        # Create response
        response = ParserResponse()
        response.set_response_type("xlsx_parse")
        response.add(result)

        return response

    def _parse(self, data: Union[str, bytes]) -> Dict[str, any]:
        """Parse XLSX and extract content.

        Args:
            data:
                XLSX file path or bytes.

        Returns:
            Dictionary with:
            - text: Markdown/HTML content
            - images: Dictionary of images
            - metadata: Document metadata
        """
        # Load workbook
        if isinstance(data, bytes):
            from io import BytesIO  # noqa: PLC0415

            workbook = load_workbook(BytesIO(data), data_only=True)
        else:
            workbook = load_workbook(data, data_only=True)

        md_content = ""
        images_dict = {}

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            md_content += f"<!-- Sheet: {sheet_name} -->\n"

            # Extract images from sheet
            for idx, image in enumerate(sheet._images):
                filename = f"{sheet_name}_image_{idx}.png"
                md_content += f"\n![Image on Sheet {sheet_name}: {idx}]({filename})\n"
                images_dict[filename] = image._data()

            # Find and extract tables
            tables = self._find_tables(sheet)
            for i, (start_row, end_row) in enumerate(tables):
                table_data = self._extract_table_data(sheet, start_row, end_row)

                # Convert table to appropriate format
                if self.table_format == "html":
                    table_str = self._table_to_html(table_data)
                else:  # markdown
                    table_str = self._table_to_markdown(table_data)

                md_content += (
                    f"\n\n<!-- Table number: {i + 1}. "
                    f"Position: ({start_row}, {end_row}) -->"
                )
                md_content += f"\n{table_str}\n"

        # Prepare images (optionally encode to base64)
        images_dict = self._prepare_images_dict(
            images_dict, encode_base64=self.encode_images_base64
        )

        # Prepare metadata
        metadata = dotdict(
            {
                "num_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "table_format": self.table_format,
            }
        )

        return {
            "text": md_content.strip(),
            "images": images_dict,
            "metadata": metadata,
        }

    def _find_tables(self, sheet) -> List[Tuple[int, int]]:
        """Find table boundaries in a sheet.

        Args:
            sheet:
                Openpyxl worksheet object.

        Returns:
            List of tuples (start_row, end_row) for each table found.
        """
        tables = []
        in_table = False
        start_row = 0

        for row in range(1, sheet.max_row + 1):
            # Check if row is empty
            row_is_empty = all(cell.value is None for cell in sheet[row])

            if not row_is_empty and not in_table:
                # Start of new table
                in_table = True
                start_row = row
            elif row_is_empty and in_table:
                # End of current table
                in_table = False
                tables.append((start_row, row - 1))

        # Add last table if incomplete
        if in_table:
            tables.append((start_row, sheet.max_row))

        return tables

    def _get_cell_value(self, sheet, row: int, col: int):
        """Get cell value handling merged cells.

        Args:
            sheet:
                Openpyxl worksheet object.
            row:
                Row number (1-indexed).
            col:
                Column number (1-indexed).

        Returns:
            Cell value, handling merged cells.
        """
        cell = sheet.cell(row, col)
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return sheet.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value

    def _extract_table_data(self, sheet, start_row: int, end_row: int) -> List[List]:
        """Extract table data from sheet.

        Args:
            sheet:
                Openpyxl worksheet object.
            start_row:
                Starting row (1-indexed).
            end_row:
                Ending row (1-indexed).

        Returns:
            2D list of table data.
        """
        # Determine maximum columns in table
        max_col = max(
            (
                cell.column
                for row in range(start_row, end_row + 1)
                for cell in sheet[row]
                if cell.value is not None
            ),
            default=1,
        )

        table_data = []
        for row in range(start_row, end_row + 1):
            row_data = [
                self._get_cell_value(sheet, row, col) for col in range(1, max_col + 1)
            ]
            table_data.append(row_data)

        return table_data

    def _table_to_markdown(self, table_data: List[List]) -> str:
        """Convert table data to Markdown format.

        Args:
            table_data:
                2D list of table data.

        Returns:
            Markdown table string.
        """
        if not table_data:
            return ""

        # First row is header
        header = table_data[0]
        rows = table_data[1:]

        # Header line
        header_line = (
            "| "
            + " | ".join(str(cell) if cell is not None else "" for cell in header)
            + " |"
        )

        # Separator line
        separator_line = "| " + " | ".join(["---"] * len(header)) + " |"

        # Data lines
        data_lines = [
            "| "
            + " | ".join(str(cell) if cell is not None else "" for cell in row)
            + " |"
            for row in rows
        ]

        return "\n".join([header_line, separator_line, *data_lines])

    def _table_to_html(self, table_data: List[List]) -> str:
        """Convert table data to HTML format.

        Args:
            table_data:
                2D list of table data.

        Returns:
            HTML table string.
        """
        if not table_data:
            return ""

        # First row is header
        header = table_data[0]
        rows = table_data[1:]

        # Start table
        html = "<table>\n"

        # Add header row
        html += "  <tr>\n"
        for cell in header:
            html += f"    <th>{str(cell) if cell is not None else ''}</th>\n"
        html += "  </tr>\n"

        # Add data rows
        for row in rows:
            html += "  <tr>\n"
            for cell in row:
                html += f"    <td>{str(cell) if cell is not None else ''}</td>\n"
            html += "  </tr>\n"

        # Close table
        html += "</table>"

        return html

    async def acall(self, data: Union[str, bytes], **kwargs) -> ParserResponse:
        """Async version of __call__. Parse an XLSX document asynchronously.

        Args:
            data:
                XLSX file path, URL, or bytes.
            **kwargs:
                Additional parsing options (currently unused).

        Returns:
            ParserResponse containing:
            - text: Markdown-formatted content with tables
            - images: Dictionary of extracted images
            - metadata: Document metadata (num_sheets, sheet_names, etc.)

        Raises:
            FileNotFoundError:
                If file path doesn't exist.
            ValueError:
                If data type is not supported or file type is invalid.
        """
        # Validate file type if it's a path
        if isinstance(data, str) and not data.startswith(("http://", "https://")):
            self._validate_file_type(data, [".xlsx", ".xlsm"])

        # Load file asynchronously if needed
        if isinstance(data, str):
            data = await self._aload_file(data)

        # Parse the document
        result = self._parse(data)

        # Create response
        response = ParserResponse()
        response.set_response_type("xlsx_parse")
        response.add(result)

        return response
